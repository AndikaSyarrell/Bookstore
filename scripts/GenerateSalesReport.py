#!/usr/bin/env python3
"""
Seller Monthly Sales Report Excel Export
Generates detailed sales report for seller dashboard
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

def create_sales_report(sales_data):
    """
    Create Excel report from sales data
    
    sales_data structure:
    {
        'seller': {'name': str, 'email': str},
        'period': {'month': int, 'year': int, 'month_name': str},
        'summary': {
            'total_orders': int,
            'total_revenue': float,
            'total_products_sold': int,
            'avg_order_value': float
        },
        'orders': [
            {
                'order_number': str,
                'date': str,
                'buyer_name': str,
                'items': int,
                'total': float,
                'status': str,
                'shipping': {
                    'carrier': str,
                    'tracking_number': str
                }
            }
        ],
        'products': [
            {
                'product_name': str,
                'sku': str,
                'quantity_sold': int,
                'revenue': float,
                'avg_price': float
            }
        ]
    }
    """
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    create_summary_sheet(wb, sales_data)
    create_orders_sheet(wb, sales_data)
    create_products_sheet(wb, sales_data)
    
    return wb

def create_summary_sheet(wb, data):
    """Create summary overview sheet"""
    sheet = wb.create_sheet('Summary', 0)
    
    # Title
    sheet['A1'] = 'SALES REPORT'
    sheet['A1'].font = Font(size=18, bold=True, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', start_color='2E75B6')
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells('A1:G1')
    sheet.row_dimensions[1].height = 30
    
    # Seller Info
    sheet['A3'] = 'Seller:'
    sheet['B3'] = data['seller']['name']
    sheet['A4'] = 'Email:'
    sheet['B4'] = data['seller']['email']
    sheet['A5'] = 'Period:'
    sheet['B5'] = f"{data['period']['month_name']} {data['period']['year']}"
    sheet['A6'] = 'Generated:'
    sheet['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    sheet['A3'].font = Font(bold=True)
    sheet['A4'].font = Font(bold=True)
    sheet['A5'].font = Font(bold=True)
    sheet['A6'].font = Font(bold=True)
    
    # Summary Stats Header
    sheet['A8'] = 'SUMMARY'
    sheet['A8'].font = Font(size=14, bold=True)
    sheet['A8'].fill = PatternFill('solid', start_color='D9E1F2')
    sheet.merge_cells('A8:G8')
    
    # Stats
    stats = [
        ('Total Orders', data['summary']['total_orders'], ''),
        ('Total Revenue', data['summary']['total_revenue'], 'Rp #,##0'),
        ('Products Sold', data['summary']['total_products_sold'], ''),
        ('Average Order Value', data['summary']['avg_order_value'], 'Rp #,##0'),
    ]
    
    row = 9
    for label, value, fmt in stats:
        sheet[f'A{row}'] = label
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'] = value
        if fmt:
            sheet[f'B{row}'].number_format = fmt
        row += 1
    
    # Column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 25

def create_orders_sheet(wb, data):
    """Create detailed orders sheet"""
    sheet = wb.create_sheet('Orders')
    
    # Headers
    headers = ['Order No', 'Date', 'Buyer', 'Items', 'Total', 'Status', 'Carrier', 'Tracking']
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='2E75B6')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 2
    for order in data['orders']:
        sheet.cell(row, 1, order['order_number'])
        sheet.cell(row, 2, order['date'])
        sheet.cell(row, 3, order['buyer_name'])
        sheet.cell(row, 4, order['items'])
        sheet.cell(row, 5, order['total']).number_format = 'Rp #,##0'
        sheet.cell(row, 6, order['status'])
        sheet.cell(row, 7, order.get('shipping', {}).get('carrier', '-'))
        sheet.cell(row, 8, order.get('shipping', {}).get('tracking_number', '-'))   
        row += 1
    
    # Totals row
    if data['orders']:
        total_row = row
        sheet.cell(total_row, 1, 'TOTAL').font = Font(bold=True)
        sheet.cell(total_row, 5, f'=SUM(E2:E{row-1})').number_format = 'Rp #,##0'
        
        for col in range(1, 10):
            sheet.cell(total_row, col).font = Font(bold=True)
            sheet.cell(total_row, col).fill = PatternFill('solid', start_color='D9E1F2')
    
    # Column widths
    widths = [15, 12, 20, 8, 15, 15, 18, 20]
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width

def create_products_sheet(wb, data):
    """Create products performance sheet"""
    sheet = wb.create_sheet('Products')
    
    # Headers
    headers = ['Product Name', 'SKU', 'Qty Sold', 'Revenue', 'Profit', 'Avg Price']
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='2E75B6')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 2
    for product in data['products']:
        sheet.cell(row, 1, product['product_name'])
        sheet.cell(row, 2, product.get('sku', '-'))
        sheet.cell(row, 3, product['quantity_sold'])
        sheet.cell(row, 4, product['revenue']).number_format = 'Rp #,##0'
        sheet.cell(row, 5, product.get('profit', 0)).number_format = 'Rp #,##0'
        sheet.cell(row, 6, product['avg_price']).number_format = 'Rp #,##0'
        row += 1
    
    # Totals row
    if data['products']:
        total_row = row
        sheet.cell(total_row, 1, 'TOTAL').font = Font(bold=True)
        sheet.cell(total_row, 3, f'=SUM(C2:C{row-1})')
        sheet.cell(total_row, 4, f'=SUM(D2:D{row-1})').number_format = 'Rp #,##0'
        sheet.cell(total_row, 5, f'=SUM(E2:E{row-1})').number_format = 'Rp #,##0'
        
        for col in range(1, 6):
            sheet.cell(total_row, col).font = Font(bold=True)
            sheet.cell(total_row, col).fill = PatternFill('solid', start_color='D9E1F2')
    
    # Column widths
    widths = [35, 15, 12, 15, 15, 15]
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: python generate_sales_report.py <json_data_file> <output_file>')
        sys.exit(1)
    
    json_file = sys.argv[1]
    output_file = sys.argv[2]
    
    with open(json_file, 'r') as f:
        sales_data = json.load(f)
    
    wb = create_sales_report(sales_data)
    wb.save(output_file)
    
    print(f'Report generated: {output_file}')